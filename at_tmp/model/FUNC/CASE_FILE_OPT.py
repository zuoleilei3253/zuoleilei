#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/7/18 15:14
# @Author  : bxf
# @File    : CASE_IMP_OPT.py
# @Software: PyCharm
import copy
import re

import tablib
import xlrd
from openpyxl import load_workbook, Workbook
from werkzeug.utils import secure_filename
from io import BytesIO
from model.FUNC.ENUM_OPT import *
from model.FUNC.CASE_INFO_OPT import *
from model.util.PUB_LOG import exeLog

PATH = getCurruntpath() + '/model/output/importfile/'
ALLOWED_EXTENSIONS = ['py', 'ico', 'js']
ALLOWED_EXCEL = ['xls', 'xlsx','xmind']


class CASE_FILE_OPT:
    '''
    用例导入
    '''

    def case_import(self, path, sheetname,group_id=None,group_id_arr=None):
        '''
        用例导入
        :param path:
        :return:
        '''
        f = xlrd.open_workbook(path)
        #sheetlist = f.sheet_names()
        # if len(sheetlist) > 0:
        #     if group_id != None:
        #         groupsql = "SELECT group_desc FROM p_group_info WHERE code='%s'" %(group_id)
        #         cursor = DB_CONN().db_Query_tuple(groupsql)
        #         grouplist=cursor.fetchall()
        #         exeLog('grouplist====================='+str(grouplist))
        #         if len(grouplist)>0:
        #             desc = grouplist[0].get('group_desc').strip()
        #             for name in sheetlist:
        #                 if name.strip().endswith(desc):
        #                     sheet = f.sheet_by_name(name)
        #                     exeLog('name====================' + name)
        #                     break
        #             if sheet:
        #                 exeLog('sheet====================is none')
        #                 sheet=f.sheet_by_index(0)
        #         else:
        #             sheet = f.sheet_by_index(0)
        #
        #     else:
        #         if not sheetname:
        #             sheet = f.sheet_by_index(0)
        #         else:
        #             sheet = f.sheet_by_name(sheetname)
        # else:
        #     sheet = f.sheet_by_index(0)
        if not sheetname:
            sheet = f.sheet_by_index(0)

        else:
            sheet = f.sheet_by_name(sheetname)

        coldict = {}
        for col in range(0, sheet.ncols):
            value = sheet.cell(0, col).value.strip()
            value = re.sub(r'\s', '', value)
            coldict[value] = col
            print(value)
        existcaseid_lists = []
        noexistcaseid_lists = []
        fail_lists=[]
        apicase_lists=[]

        for row in range(1, sheet.nrows):
            case_list = dict()
            if group_id != None:
                case_list['group_id'] = group_id
                case_list['group_id_arr'] = json.dumps(group_id_arr)
            else:
                case_list['group_id'] = ''
                case_list['group_id_arr'] = json.dumps([])
            case_id=sheet.cell(row, coldict.get(ENUM_OPT('case_title_list').get_val('id'))).value
            if case_id is None or  case_id=='':
                case_id = newID().CS_ID()
                nocaseid = True
            else:
                nocaseid = False
            try:
                case_list['case_id'] = case_id.strip()  # 用例ID
                case_list['case_path'] = sheet.cell(row, coldict.get( ENUM_OPT('case_title_list').get_val('casedir'))).value.strip().replace('\'','\'\'')  # 用例目录
                case_list['case_desc'] = sheet.cell(row, coldict.get(ENUM_OPT('case_title_list').get_val('casename'))).value.strip().replace('\'','\'\'')  # 用例名称
                case_list['rqmt_id'] = sheet.cell(row, coldict.get(ENUM_OPT('case_title_list').get_val('requestid'))).value.strip().replace('\'','\'\'')  # 需求ID
                case_list['case_init'] = sheet.cell(row, coldict.get(ENUM_OPT('case_title_list').get_val('caseinit'))).value.strip().replace('\'','\'\'')  # 前置条件
                case_list['case_step '] = sheet.cell(row, coldict.get(ENUM_OPT('case_title_list').get_val('casestep'))).value.strip().replace('\'','\'\'')  # 用例步骤
                case_list['case_prev_data'] = sheet.cell(row, coldict.get(ENUM_OPT('case_title_list').get_val('caseresult'))).value.strip().replace('\'','\'\'')  # 预期结果
                case_type_int = sheet.cell(row, coldict.get(ENUM_OPT('case_title_list').get_val('casetype'))).value.strip().replace('\'','\'\'')  # 用例类型
                case_list['case_type'] = ENUM_OPT('case_type').get_val(case_type_int)
                case_exe_status_int =sheet.cell(row, coldict.get(ENUM_OPT('case_title_list').get_val('casestatus'))).value.strip().replace('\'','\'\'')  # 用例状态

                case_list['case_exe_status'] = ENUM_OPT('case_exe_status').get_val(case_exe_status_int)
                case_level_int =sheet.cell(row, coldict.get(ENUM_OPT('case_title_list').get_val('caselevel'))).value.strip().replace('\'','\'\'') # 用例等级
                case_list['case_level'] = ENUM_OPT('case_level').get_val(case_level_int)
                case_list['case_builder'] =sheet.cell(row, coldict.get(ENUM_OPT('case_title_list').get_val('casebuild'))).value.strip().replace('\'','\'\'') # 创建人
                case_exe_env_int = sheet.cell(row, coldict.get(ENUM_OPT('case_title_list').get_val('casedev'))).value.strip().replace('\'','\'\'') # 环境类型
                case_list['case_exe_env'] = ENUM_OPT('case_exe_env').get_val(case_exe_env_int)
                case_exe_type_int = sheet.cell(row, coldict.get(ENUM_OPT('case_title_list').get_val('exetype'))).value.strip().replace('\'','\'\'') # 执行类型
                case_list['case_exe_type'] = ENUM_OPT('case_exe_type').get_val(case_exe_type_int)
                case_exe_plugin_int = sheet.cell(row, coldict.get(ENUM_OPT('case_title_list').get_val('plug'))).value.strip().replace('\'','\'\'') # 执行插件
                case_list['case_exe_plugin '] = ENUM_OPT('case_exe_plugin').get_val(case_exe_plugin_int)
                apiurl=coldict.get(ENUM_OPT('case_title_list').get_val('apiurl'))
                if apiurl:
                    apicase={}
                    api_url = sheet.cell(row, apiurl).value.strip().replace('\'', '\'\'')  #接口url
                    api_method=sheet.cell(row, coldict.get(ENUM_OPT('case_title_list').get_val('method'))).value.strip().replace('\'', '\'\'').upper()  #接口请求方法
                    api_params=sheet.cell(row,coldict.get(ENUM_OPT('case_title_list').get_val('apiparams'))).value.strip().replace('\'', '\'\'')  #接口参数
                    api_check=sheet.cell(row,coldict.get(ENUM_OPT('case_title_list').get_val('apicheck'))).value.strip().replace('\'', '\'\'')  #接口检查点
                    apicase['case_id'] = case_id.strip()
                    apicase['api_url']=api_url
                    apicase['api_method']=api_method
                    apicase['api_params']=api_params
                    apicase['api_check']=api_check
                    apicase_lists.append(apicase)
                case_list['row']=row
            except Exception as e:
                exeLog("*******读取用例失败 跳过继续" + str(e))
                fail_lists.append(row+1)
                continue

            if nocaseid:
                noexistcaseid_lists.append(case_list)
            else:
                existcaseid_lists.append(case_list)
        return noexistcaseid_lists, existcaseid_lists,fail_lists,apicase_lists
        # wb = load_workbook(path)
        # if sheetname == "":
        #     ws = wb.active
        # else:
        #     ws = wb[sheetname]
        # case_lists = []
        # for j in range(1, ws.max_row):
        #     case_list = dict()
        #     if group_id !=None:
        #         case_list['group_id']=group_id
        #         case_list['group_id_arr'] = json.dumps(group_id_arr)
        #
        #     else:
        #         case_list['group_id']=''
        #         case_list['group_id_arr'] = json.dumps([])
        #     if ws[j + 1][0].value is None:
        #         case_id = newID().CS_ID()
        #     else:
        #         case_id=ws[j + 1][0].value
        #     case_list['case_id']=case_id #用例ID
        #     case_list['case_path'] = ws[j + 1][1].value  # 用例目录
        #     case_list['case_desc'] = ws[j + 1][2].value  # 用例名称
        #     # case_list['rqmt_id'] = ws[j + 1][3].value  # 需求ID
        #     case_list['case_init'] = ws[j + 1][4].value  # 前置条件
        #     case_list['case_step '] = ws[j + 1][5].value  # 用例步骤
        #     case_list['case_prev_data'] = ws[j + 1][6].value  # 预期结果
        #     case_type_int = ws[j + 1][7].value  # 用例类型
        #     case_list['case_type'] = ENUM_OPT('case_type').get_val(case_type_int)
        #     case_exe_status_int = ws[j + 1][8].value  # 用例状态
        #     case_list['case_exe_status'] = ENUM_OPT('case_exe_status').get_val(case_exe_status_int)
        #     case_level_int = ws[j + 1][9].value  # 用例等级
        #     case_list['case_level'] = ENUM_OPT('case_level').get_val(case_level_int)
        #     case_list['case_builder'] = ws[j + 1][10].value  # 创建人
        #     case_exe_env_int = ws[j + 1][11].value  # 环境类型
        #     case_list['case_exe_env'] = ENUM_OPT('case_exe_env').get_val(case_exe_env_int)
        #     case_exe_type_int = ws[j + 1][12].value  # 执行类型
        #     case_list['case_exe_type'] = ENUM_OPT('case_exe_type').get_val(case_exe_type_int)
        #     case_exe_plugin_int = ws[j + 1][13].value  # 执行插件
        #     case_list['case_exe_plugin '] = ENUM_OPT('case_exe_plugin').get_val(case_exe_plugin_int)
        #     case_lists.append(case_list)

        #return case_lists


    def insert_table(self, table, path, sheetname, rqmt_id=None,group_id=None,group_id_arr=None):
        '''
        将数据插入到数据库
        :param table: 表名
        :param path: 文件路径
        :param sheetname: 表单名
        :return: 插入成功数
        '''
        noexistcaseid_lists, existcaseid_lists,fail_lists,apicase_lists = self.case_import(path, sheetname,group_id,group_id_arr)
        insert_counts = int()
        fail_count=int()
        message=''
        isfail=False
        if len(noexistcaseid_lists)>0:
            for i in noexistcaseid_lists:
                case_id=i.get('case_id')
                try:
                    row=i.get('row')
                    del i['row']
                    counts = insertToDatabase(table, i, rqmt_id=rqmt_id)
                    insert_counts = insert_counts + counts
                except Exception as e:
                    isfail=True
                    exeLog("*******导入失败跳过继续+++错误信息为" + str(e))
                    fail_count+=1
                    fail_lists.append(row+1)
                    continue
                if apicase_lists:
                    for case in apicase_lists:
                        if case.get('case_id')==case_id and case.get('api_url'):
                            apicase_step=self.apicasestep(case,group_id)
                            SUITE_OPT().suitcase(case_id,json.dumps(apicase_step))
                            break

        if len(existcaseid_lists)>0:
            for i in existcaseid_lists:
                case_id = i.get('case_id')
                try:
                    row = i.get('row')
                    del i['row']
                    selectsql= "SELECT * FROM %s WHERE case_id='%s'" % (table,case_id)
                    count =  DB_CONN().db_Update(selectsql)
                    if count>0:
                        update_counts=updateToDatabase(table, i, case_id=case_id)
                        insert_counts+=update_counts
                    else:
                        isfail = True
                        message='已存在的用例id在数据库中不存在，更新失败'
                        fail_lists.append(row+1)
                except Exception as e:
                    fail_count+=1
                    isfail = True
                    exeLog("*******更新失败跳过继续+++错误信息为" + str(e))
                    fail_lists.append(row+1)
                    continue
            if apicase_lists:
                for case in apicase_lists:
                    if case.get('case_id') == case_id and case.get('api_url'):
                        apicase_step = self.apicasestep(case, group_id)
                        SUITE_OPT().suitcase(case_id, json.dumps(apicase_step))
                        break
        if not isfail and len(fail_lists)==0:
           return '导入成功:' + str(insert_counts) + '条\n'+' 失败:0条'
        else:
            return '导入成功:' + str(insert_counts) + '条\n'+' 失败:'+str(len(fail_lists))+'条\n'+message+'\n 失败行数为:'+str(fail_lists)
        #     #return insert_counts
        # except Exception as e:
        #     exeLog("*******导入失败+++错误信息为" + str(e))
        #     #return False
        #     return '导入失败，请检查导入文件的数据内容是否正确！~'
    def apicasestep(self,case,group_id):
        api_url = case.get('api_url')
        suite_case={}
        suite_data = []
        if case.get('api_url'):
            sql = "SELECT api_id FROM api_case_info WHERE uri=%s AND method=%s AND group_id=%s"
            params = (api_url, case.get('api_method'), group_id)
            cur = DB_CONN().db_Query_Json(sql, params)
            if cur:
                api_list = cur.fetchall()
                if len(api_list) == 1:
                    step_id="ST-0001"
                    info_id=api_list[0].get('api_id')
                    api_params=case.get('api_params')
                    if api_params:
                        paramsdict=json.loads(api_params)
                    else:
                        paramsdict={}
                    init_data=self.initparams(paramsdict)
                    tail_data=self.tailparams(json.loads(case.get('api_check')))
                    casesuite={}
                    casesuite['step_id']=step_id
                    casesuite['info_id']=info_id
                    casesuite['init_data']=init_data
                    casesuite['tail_data']=tail_data
                    suite_data.append(casesuite)
        suite_case['suite_data']=suite_data
        return suite_case
    def initparams(self,paramsdict):
        i=0
        init_data=[]
        for key, value in paramsdict.items():
            key_data=[]
            init = {}
            init["arr_index"] = i
            init["custom_val"] = value
            init["key"] = key
            init["assign_type"] = "custom"
            i += 1
            key_data.append(key)
            key_data.append('手工输入: '+str(value))
            key_data.append( init)
            init_data.append(key_data)
        return init_data
    def tailparams(self,valuedict):
        tail_data = []
        init={
            'check_id': 'CH-005',
            'check_desc': '返回内容匹配JSON（部分）',
            'judge_type': 'json_part',
            'input_info': '输入数据为要匹配的json格式片段<br>【输入示例】：{"code":200} ',
            'cp_id': 'CP-0001',
            'value': {
                'msg': 200
            }
        }
        init['value']=valuedict
        tail_data.append(init)
        return tail_data


    def build_xls(self, headers, data):
        '''
        生成导出方法
        :param headers:
        :param table:
        :return:databook
        '''


        if data:
            data = list(data)
        else:
            data = [()]
        data_set = tablib.Dataset(*data, headers=headers)
        data_book = tablib.Databook()
        data_book.add_sheet(data_set)
        return data_book

    def allowed_file(self, filename):
        '''
        判断是否允许的文件类型
        :return:
        '''
        return '.' in filename and \
               filename.rsplit('.', 1)[1] in ALLOWED_EXCEL

    def imt_excel(self, opt_files, table, rqmt_id,group_id,group_id_arr):
        if opt_files and self.allowed_file(opt_files.filename):
            opt_filename = secure_filename(opt_files.filename)
            filenames=''
            if rqmt_id:
                filenames = rqmt_id + '_' + str(newID().nomal_time()) + '.xlsx'
            else:
                filenames = str(newID().nomal_time()) + '.xlsx'
            opt_files.save(os.path.join(PATH, filenames))
            path = PATH + filenames

            sheetname = ''
            result = self.insert_table(table, path, sheetname, rqmt_id,group_id,group_id_arr)
            # if result:
            #     # self.case_import(path, sheetname)
            #     # apiEXCEL(path, sheetname)
            #     return_data = respdata().sucessMessage('', '导入成功！~导入用例：' + str(result) + '条')
            #     return json.dumps(return_data, ensure_ascii=False)
            # else:
            return_data = respdata().sucessMessage('', result)
            return json.dumps(return_data, ensure_ascii=False)
        else:
            return_data = respdata().failMessage('', '导入失败，请检查导入文件类型！~')
            return json.dumps(return_data, ensure_ascii=False)

    def imt_xmind(self, opt_files):
        if opt_files and self.allowed_file(opt_files.filename):
            print("_________________")
            filenames = opt_files.filename
            print(filenames)
            opt_files.save(os.path.join(PATH, filenames))
            return PATH + filenames
        else:
            return_data = respdata().failMessage('', '导入失败，请检查导入文件类型！~')
            return json.dumps(return_data, ensure_ascii=False)

    def rqmt_case_export_opt(self):
        '''
        案例导出
        :param data:
        :return:
        '''
        headers = (
            '测试项目组', '接口人', '总用例数', '自动化用例数', '测试覆盖率', '测试执行率', '问题发现率', '统计时间')  # headers 为表头的字段名称，格式为（'字段1','字段2','字段3'）
        sql = "SELECT total ,a.group_id,b.num atnum ,(b.num/a.total) AS cover_rate,IF(c.num IS NULL,0,(c.num/a.total)) AS exe_rate,IF(d.num IS NULL,0,(d.num/a.total)) AS bug_rate FROM (SELECT group_id,COUNT(1) AS total FROM regress_case_info WHERE 1=1 GROUP BY group_id) a LEFT JOIN (SELECT group_id,COUNT(1) AS num FROM regress_case_info WHERE case_exe_type!=1 GROUP BY group_id) b ON b.group_id=a.group_id LEFT JOIN (SELECT group_id,COUNT(1) AS num FROM regress_case_result x1 INNER JOIN regress_case_info x2 ON x2.case_id=x1.case_id WHERE batch_id=(SELECT MAX(batch_id) FROM regress_case_result) GROUP BY group_id) c ON c.group_id=a.group_id LEFT JOIN (SELECT group_id,COUNT(1) AS num FROM regress_case_result x1 INNER JOIN regress_case_info x2 ON x2.case_id=x1.case_id WHERE case_result=3 GROUP BY group_id) d ON d.group_id=a.group_id "
        sql_all="SELECT total,b.num atnum,(b.num/a.total) AS cover_rate,(c.num/a.total) AS exe_rate,(d.num/a.total) AS bug_rate FROM (SELECT '1' AS id,COUNT(1) AS total FROM regress_case_info WHERE 1=1) a INNER JOIN (SELECT '1' AS id,COUNT(1) AS num FROM regress_case_info WHERE case_exe_type!=1) b ON b.id=a.id INNER JOIN (SELECT '1' AS id,COUNT(1) AS num FROM regress_case_result WHERE batch_id=(SELECT MAX(batch_id) FROM regress_case_result)) c ON c.id=a.id INNER JOIN (SELECT '1' AS id,COUNT(1) AS num FROM regress_case_result WHERE case_result=3) d ON d.id=a.id"# 为要导出的sql
        get_data=getTupleFromDatabase(sql)
        get_data2=getTupleFromDatabase(sql_all)
        get_data.append(get_data2[0])
        databook = CASE_FILE_OPT().build_xls(headers, get_data)
        output = BytesIO()
        output.write(databook.xls)
        return_data = output.getvalue()
        return return_data

    def case_write(self, filename, table, group_id=None,**kwargs):
        sql_doc = ''
        wb = load_workbook(PATH + 'case.xlsx')
        ws = wb['sheet1']
        if group_id:
            sql_doc=" where group_id like '"+str(group_id)+"%'"
        else:
            for i in kwargs:
                col = i
                val = kwargs[i]
                sql_doc = ' WHERE ' + col + '="' + str(val) + '"'
        data_sql = "select case_id,case_path,case_desc,rqmt_id,case_init,case_step ,case_prev_data,case_type,case_exe_status,case_level,case_builder,case_exe_env,case_exe_type,case_exe_plugin  from " + table + sql_doc

        case_lists = getJsonFromDatabase(data_sql)
        if case_lists:
            case_alls = []
            for i in case_lists:
                casel_list = []
                case_id = i['case_id']
                casel_list.append(case_id)
                case_path = i['case_path']
                casel_list.append(case_path)
                case_desc = i['case_desc']
                casel_list.append(case_desc)
                rqmt_id = i['rqmt_id']
                casel_list.append(rqmt_id)
                case_init = i['case_init']
                casel_list.append(case_init)
                case_step = i['case_step']
                casel_list.append(case_step)
                case_prev_data = i['case_prev_data']
                casel_list.append(case_prev_data)
                case_type_int = i['case_type']
                case_type = ENUM_OPT('case_type').get_key(case_type_int)
                casel_list.append(case_type)
                case_exe_status_int = i['case_exe_status']
                case_exe_status = ENUM_OPT('case_exe_status').get_key(case_exe_status_int)
                casel_list.append(case_exe_status)
                case_level_int = i['case_level']
                case_level = ENUM_OPT('case_level').get_key(case_level_int)
                casel_list.append(case_level)
                case_builder = i['case_builder']
                casel_list.append(case_builder)
                case_exe_env_int = i['case_exe_env']
                case_exe_env = ENUM_OPT('case_exe_env').get_key(case_exe_env_int)
                casel_list.append(case_exe_env)
                case_exe_type_int = i['case_exe_type']
                case_exe_type = ENUM_OPT('case_exe_type').get_key(case_exe_type_int)
                casel_list.append(case_exe_type)
                case_exe_plugin_int = i['case_exe_plugin']
                case_exe_plugin = ENUM_OPT('case_exe_plugin').get_key(case_exe_plugin_int)
                casel_list.append(case_exe_plugin)
                case_alls.append(casel_list)
            for j in range(len(case_alls)):
                for g in range(len(case_alls[0])):
                    ws.cell(row=j + 2, column=g + 1, value=case_alls[j][g])
        else:
            ws.cell(row=2, column=1, value='该需求下无用例！~，请检查！~~')

        wb.save(filename)

    def api_write(self, filename, groupId):
        '''
        接口信息导出功能
        :param filename:
        :param groupId:
        :return:
        '''
        sql_doc = ''
        wb = load_workbook(PATH + 'api_info.xlsx')
        ws = wb['Sheet1']
        data_sql = "select api_id,uri,title,method,params,headers,return_1,return_2,api_type,api_need_login,api_is_remold,api_status,api_path from api_case_info where group_id like '"+str(groupId)+"%'"
        case_lists = getJsonFromDatabase(data_sql)
        if case_lists:
            case_alls = []
            for i in case_lists:
                casel_list = []
                api_id = i['api_id']
                casel_list.append(api_id)
                uri = i['uri']
                casel_list.append(uri)
                title = i['title']
                casel_list.append(title)
                method = i['method']
                casel_list.append(method)
                params = i['params']
                casel_list.append(params)
                headers = i['headers']
                casel_list.append(headers)
                return_1 = i['return_1']
                casel_list.append(return_1)
                return_2 = i['return_2']
                casel_list.append(return_2)
                api_type=i['api_type']
                api_type = ENUM_OPT('api_type').get_key(api_type)
                casel_list.append(api_type)
                api_need_login = i['api_need_login']
                api_need_login = ENUM_OPT('api_need_login').get_key(api_need_login)
                casel_list.append(api_need_login)
                api_is_remold = i['api_is_remold']
                api_is_remold = ENUM_OPT('api_is_remold').get_key(api_is_remold)
                casel_list.append(api_is_remold)
                api_status = i['api_status']
                api_status = ENUM_OPT('api_status').get_key(api_status)
                casel_list.append(api_status)
                api_path = i['api_path']
                casel_list.append(api_path)
                case_alls.append(casel_list)
            for j in range(len(case_alls)):
                for g in range(len(case_alls[0])):
                    ws.cell(row=j + 4, column=g + 1, value=case_alls[j][g])
        else:
            ws.cell(row=4, column=1, value='该需求下无用例！~，请检查！~~')

        wb.save(filename)
