# -*- coding: utf-8 -*-

from model.util.Third_XMIND import xmind
# import tkFileDialog
import openpyxl
import os
import json
import time
from model.util.PUB_LOG import *
testcase = []
testcasedir = 0
testcasename = 0
i = 1
k = 0


def translate_to_excel(xmind_file,template_path,sasve_path):
    # 写入内容
    def write(testcase):
        exeLog("用例：" + str(testcase))
        #global 的作用是将变量声明为全局变量。
        global i
        i += 1
        for a in range(1,len(testcase)+1):
            # #第i行，第testcasename 用例名称，赋值用例名称
            # if "：" in str(testcase[a-1]):
            #     ws.cell(row=i, column=a, value=testcase[a-1].split("：")[1])
            #     if i < 10 and a == 1:
            #         ws.cell(row=i, column=1, value=str(testcase[0]) + "_00" + str(i-1))
            #     if 10 < i < 100 and a == 1:
            #         ws.cell(row=i, column=1, value=str(testcase[0]) + "_0" + str(i - 1))
            #     if a == 13:
            #         ws.cell(row=i, column=14, value=testcase[a].split("：")[1])
            # else:
            ws.cell(row=i, column=a, value=testcase[a - 1])
            if i < 10 and a == 1 or i == 10 and a == 1:
                ws.cell(row=i, column=1, value=str(testcase[0]) + "_00" + str(i - 1))
            if 10 < i < 100 and a == 1:
                ws.cell(row=i, column=1, value=str(testcase[0]) + "_0" + str(i - 1))
            # if a == 2:
            ws.cell(row=i, column=2, value=str(testcase[1]))

            # if a == 3:
            ws.cell(row=i, column=3, value=str(testcase[2]))
            # if a == 4:
            ws.cell(row=i, column=4, value="无")
            if len(testcase) < 13:
                ws.cell(row=i, column=5, value="无")
                ws.cell(row=i, column=6, value=str(testcase[3]))
                ws.cell(row=i, column=7, value=str(testcase[4]))
                ws.cell(row=i, column=8, value=str(testcase[5]))
                ws.cell(row=i, column=9, value=str(testcase[6]))
                ws.cell(row=i, column=10, value=str(testcase[7]))
                ws.cell(row=i, column=11, value=str(testcase[8]))
                ws.cell(row=i, column=12, value=str(testcase[9]))
                ws.cell(row=i, column=13, value=str(testcase[10]))
                ws.cell(row=i, column=14, value=str(testcase[11]))
            else:
                ws.cell(row=i, column=5, value=str(testcase[3]))
                ws.cell(row=i, column=6, value=str(testcase[4]))
                ws.cell(row=i, column=7, value=str(testcase[5]))
                ws.cell(row=i, column=8, value=str(testcase[6]))
                ws.cell(row=i, column=9, value=str(testcase[7]))
                ws.cell(row=i, column=10, value=str(testcase[8]))
                ws.cell(row=i, column=11, value=str(testcase[9]))
                ws.cell(row=i, column=12, value=str(testcase[10]))
                ws.cell(row=i, column=13, value=str(testcase[11]))
                ws.cell(row=i, column=14, value=str(testcase[12]))
                # if a == 13:
                #     ws.cell(row=i, column=14, value=testcase[a])
    #读ximd，转excel
    def search(topic):
        global testcase
        topics=topic.getSubTopics()
        if isinstance(topics,list):
            for subTopic in topics:
                title = subTopic.getTitle()
                testcase.append(title)                
                notes = subTopic.getNotes()
                if not notes is None:
                    content = notes.getContent()   
                result = search(subTopic)
                if result is None:
                    testcase = testcase[:-1]
        else:
            write(testcase)
#logger.info("读 xmind写入excel：")
        #logger.info(json.dumps(testcase, ensure_ascii=False))
            json.dumps(testcase, ensure_ascii=False)
    def getcol():
    #判断列
        #声明全局变量
        global testcasedir, testcasename, testcaseresult, testid, task_id, preconditon, caseflow, casetype, casestatus,caselevel,creator,envtype,exetype,exeplus
        #maxcol在下面已经获取，最大列数
        for cx in range(1,maxcol+1):
            #第一行，第cx列的值赋给txt
            txt = ws.cell(row=1, column=cx).value
            # print('txt:'+str(txt))
            if txt == None:
                #logger.info("TXT is wrong")
                exeLog('Excel模板错误，第一行获取值为None')
            else:
                #如果第一行的列值为“用例目录，那么将列数赋值给testcasedir
                if txt == "用例目录":
                    testcasedir = cx
                    #logger.info("用例目录在第 %d 列", testcasedir)
                    # print("用例目录在第 %d 列", testcasedir)
                # 第一行的列值为“用例名称，那么将列数赋值给testcasename
                if txt == "用例名称":
                    testcasename = cx
                    #logger.info("用例名称在第 %d 列", testcasename)
                    # print ("用例名称在第 %d 列", testcasename)
                # 第一行的列值为“预期结果，那么将列数赋值给testcaseresult
                if txt == "预期结果":
                    testcaseresult = cx
                    #logger.info("预期结果在第 %d 列", testcasename)
                    # print ("预期结果在第 %d 列", testcaseresult)
                if txt == "用例ID":
                    testid = cx
                    # print("用例ID在第 %d 列", testid)
                if txt == "需求ID":
                    task_id = cx
                    # print("需求ID在第 %d 列", task_id)
                if txt == "前置条件":
                    preconditon = cx
                if txt == "用例步骤":
                    caseflow = cx
                if txt == "用例类型":
                    casetype = cx
                if txt == "用例状态":
                    casestatus = cx
                if txt == "用例等级":
                    caselevel = cx
                if txt == "创建人":
                    creator = cx
                if txt == "环境类型":
                    envtype = cx
                if txt == "执行类型":
                    exetype = cx
                if txt == "执行插件":
                    exeplus = cx
        if testcasedir and testcasename and testcaseresult and testid and task_id and preconditon and caseflow and casetype \
                and casestatus and caselevel and creator and envtype and exetype and exeplus:
            #调用search(topic)方法
            search(topic)
        else:
            #logger.error("Excel is wrong")
            exeLog("Excel模板错误，字段缺少")
    try:
    #打开xmind
    #xmindfilename=tkFileDialog.askopenfilename(**file_opt)
        #声明全局变量
        global testcase,testcasedir,testcasename, testcaseresult,i,k,content
        mindMap = xmind.load(xmind_file)
    except Exception:
        #logger.error("Error,please input a excel file")
        exeLog("上传的文件格式错误，请上传Excel文件")
    else:
        sheet = mindMap.getPrimarySheet()
        RootTopic = sheet.getRootTopic()
        topic = RootTopic.getTopics()
        testcase.append(RootTopic.getTitle())
        # content = testcase[3]
            #打开excel模板
        try:
            #excel引入库
            # wb = openpyxl.load_workbook("D:\\fwx\\Excel_ximind_tool\\templates.xlsx")
            wb = openpyxl.load_workbook(template_path)
            exeLog('=============导入Excel模板成功=============')
        except:
            #logger.error("Error!templates.xlsx does not exist")
            exeLog("Excel模板不存在")
        else:
            # 打开一个xlsx文件
            #指定sheet页(第一页)
            ws = wb.worksheets[0]
            #赋值最大列数
            maxcol = ws.max_column
            #调用def getcol():方法
            getcol()
            # print ('44444444444444444')
            #另保存文件
            #赋值文件名
            # save_file_name=xmind_file.split("/")[-1].split(".")[0] + '_' + str(time.strftime("%Y%m%d-%H-%M-%S", time.localtime())) + ".xlsx"
            #保存一个新的excel.xlsx
            wb.save(sasve_path)
            #logger.info('Succeed!')
            exeLog('=============Xmind转Excel表格成功！=============')

# #初始化数据
#             testcase = []
#             testcasedir = 0
#             testcasename = 0
#             i = 1
#             k = 0
#             return os.path.abspath(save_file_name)


if __name__ == '__main__':
    # xmind_file='D:\\fwx\\tuandai\\经理\\用例.xmind'
    template_file = "F:\\AT_TIMETASK-update\\Xmind_to_Excel\\用例模板_final.xlsx"
    xmind_file = 'F:\\AT_TIMETASK-update\\Xmind_to_Excel\\测试用例.xmind'
    sasve_path = "./"
    translate_to_excel(xmind_file,template_file,sasve_path)