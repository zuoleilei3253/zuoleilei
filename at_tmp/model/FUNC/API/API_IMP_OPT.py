#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/11/20 15:05
# @Author  : bxf
# @File    : API_IMP_OPT.py
# @Software: PyCharm

from model.util.PUB_LOG import *
from openpyxl import load_workbook
from model.util.newID import *
from model.util.PUB_RESP import *
from model.util.PUB_DATABASEOPT import *
from model.FUNC.GROUP_OPT import *


PATH=getCurruntpath()+'/model/output/importfile/'
ALLOWED_EXTENSIONS=['py', 'ico', 'js']
ALLOWED_EXCEL=['xls','xlsx']


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1] in ALLOWED_EXCEL

# api excel导入
def apiIMP(opt_files,group_id):

    if opt_files and allowed_file(opt_files.filename):

        opt_files.save(os.path.join(PATH, opt_files.filename))
        path = PATH + opt_files.filename
        sheetname = ''
        apiEXCEL(path, sheetname, group_id)
        return_data=respdata().sucessResp('')
        return json.dumps(return_data, ensure_ascii=False)
    else:
        a = dict()
        a['response'] = '请求错误，错误信息为： 上传文件类型错误！ '
        return json.dumps(resp(201, '上传文件类型错误或其他原因', a), ensure_ascii=False)



def apiEXCEL(path, sheetname, groupid):
    try:
        exeLog("开始导入接口信息")
        db = getJsonMysql()
        print("排查错误")
        sucess_num = 0
        fail_num = 0
        surplus = 0
        wb = load_workbook(path)
        if sheetname == "":
            ws = wb.active
        else:
            ws = wb[sheetname]
        sql = 'select api_id from api_case_info'
        result = db.get_data(sql)
        api_list = loop_api(result)
        for j in range(3, ws.max_row):
            try:
                api_id = ws[j + 1][0].value
                if ws[j + 1][2].value is None:
                    surplus = surplus + 1
                if api_id is None:
                    try:
                        empty_api(ws, j, groupid, db)
                        sucess_num = sucess_num + 1
                    except Exception as msg:
                        fail_num = fail_num + 1
                        errorLog(str(msg))
                elif api_id in api_list:
                    try:
                        existence_api(ws, j, groupid, db, api_id)
                        sucess_num = sucess_num + 1
                    except Exception as msg:
                        fail_num = fail_num + 1
                        errorLog(str(msg))
                elif api_id not in api_list:
                    nonexistent_api()
                    fail_num = fail_num + 1
            except Exception as msg:
                errorLog(str(msg))

        message = "导入完成，其中导入成功： " + str(sucess_num) + " 条，失败：" + str(fail_num - surplus) + " 条，请检查是否存在基本信息！"
        return_data = respdata().sucessMessage('', message)
        dataOptLog("接口信息导入结束")
        return json.dumps(return_data, ensure_ascii=False)
    except Exception as msg:
        errorLog(str(msg))
        return_data = respdata().failMessage('', '导入失败' + str(msg))
        return json.dumps(return_data, ensure_ascii=False)

def empty_api(ws, j, groupid, db):
    api_uri = ws[j + 1][1].value
    if api_uri != None:
        print("测试一下")
        case_desc = ws[j + 1][2].value
        method = ws[j + 1][3].value
        data = ws[j + 1][4].value
        headers = ws[j + 1][5].value
        return1 = ws[j + 1][6].value
        return2 = ws[j + 1][7].value
        api_type = ws[j + 1][8].value
        api_path=ws[j+1][12].value

        # 接口類型
        if api_type is None:
            api_type = None
        elif "查" in api_type:
            api_type = 1
        elif "新" in api_type:
            api_type = 2
        elif "改" in api_type:
            api_type = 3
        elif "刪" in api_type:
            api_type = 4

        # 接口登錄態
        api_need_login = ws[j + 1][9].value
        if api_need_login is None:
            api_need_login = api_need_login
        elif "否" in api_need_login:
            api_need_login = 0
        elif "是" in api_need_login:
            api_need_login = 1

        # 接口是否重構
        api_is_remold = ws[j + 1][10].value
        if api_is_remold is None:
            api_is_remold = None
        elif "否" in api_is_remold:
            api_is_remold = 0
        elif "是" in api_is_remold:
            api_is_remold = 1

        # 接口狀態
        api_status = ws[j + 1][11].value
        if api_status is None:
            api_status = None
        elif "无" in api_status:
            api_status = 0
        elif "有" in api_status:
            api_status = 1
        elif '改' in api_status:
            api_status=2
        id = newID().apiId()
        group_id = groupid
        if data != None:
            init_pa = changRaw(json.loads(data), 'api')
            sql = 'insert into api_case_info (api_id,uri,method,title,headers,params,init_data,group_id,api_type,api_need_login,api_is_remold,api_status,api_path) VALUE (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
            aparams = (
                id, api_uri, method, case_desc, headers, data, json.dumps(init_pa), group_id, api_type, api_need_login, api_is_remold, api_status,api_path)
        else:
            sql = 'insert into api_case_info (api_id,uri,method,title,headers,params,group_id,api_type,api_need_login,api_is_remold,api_status,api_path) VALUE (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
            aparams = (id, api_uri, method, case_desc, headers, data, group_id, api_type, api_need_login, api_is_remold, api_status,api_path)

        db.exeUpdateByParamJson(sql, aparams)
        return_data = respdata().sucessMessage('', '插入成功')
        print(return_data)
        return json.dumps(return_data, ensure_ascii=False)


def existence_api(ws, j, groupid, db, api_id):
    api_uri = ws[j + 1][1].value
    if api_uri != None:
        print("测试一下")
        case_desc = ws[j + 1][2].value
        method = ws[j + 1][3].value
        data = ws[j + 1][4].value
        headers = ws[j + 1][5].value
        return1 = ws[j + 1][6].value
        return2 = ws[j + 1][7].value
        api_type = ws[j + 1][8].value
        api_path=ws[j+1][12].value

        # 接口類型
        if api_type is None:
            api_type = None
        elif "查" in api_type:
            api_type = 1
        elif "新" in api_type:
            api_type = 2
        elif "改" in api_type:
            api_type = 3
        elif "刪" in api_type:
            api_type = 4

        # 接口登錄態
        api_need_login = ws[j + 1][9].value
        if api_need_login is None:
            api_need_login = api_need_login
        elif "否" in api_need_login:
            api_need_login = 0
        elif "是" in api_need_login:
            api_need_login = 1

        # 接口是否重構
        api_is_remold = ws[j + 1][10].value
        if api_is_remold is None:
            api_is_remold = None
        elif "否" in api_is_remold:
            api_is_remold = 0
        elif "是" in api_is_remold:
            api_is_remold = 1

        # 接口狀態
        api_status = ws[j + 1][11].value
        if api_status is None:
            api_status = None
        elif "无" in api_status:
            api_status = 0
        elif "有" in api_status:
            api_status = 1
        elif '改' in api_status:
            api_status = 2

        group_id = groupid
        if data != None:
            init_pa = changRaw(json.loads(data), 'api')
            # sql = "update api_case_info set api_path ='" + api_path + "',uri='" + api_uri + "',method='" + method + "',title='" + case_desc + "',headers='" + headers + "',params='" + data + "',init_data='" + json.dumps(
            #     init_pa) + "',group_id='" + str(group_id) + "',api_type='" + str(api_type) + "',api_need_login='" + str(
            #     api_need_login) + "',api_is_remold='" + str(api_is_remold) + "',api_status='" + str(
            #     api_status) + "' where api_id ='" + api_id + "'"
            sql = 'update api_case_info set api_path =%s,uri=%s,method=%s,title=%s,headers=%s,params=%s,init_data=%s,group_id=%s,api_type=%s,api_need_login=%s,api_is_remold=%s,api_status=%s where api_id =%s'
            params=(api_path,api_uri,method,case_desc,headers,data,json.dumps(init_pa),str(group_id),str(api_type),str(api_need_login),str(api_is_remold),str(api_status),api_id)
        else:

            # sql = "update api_case_info set api_path ='" + api_path + "',uri='" + api_uri + "',method='" + method + "',title='" + case_desc + "',headers='" + headers + "',params='" + data +  "',group_id='" + str(group_id) + "',api_type='" + str(api_type) + "',api_need_login='" + str(
            #     api_need_login) + "',api_is_remold='" + str(api_is_remold) + "',api_status='" + str(
            #     api_status) + "' where api_id ='" + api_id + "'"
            sql = 'update api_case_info set api_path =%s,uri=%s,method=%s,title=%s,headers=%s,params=%s,group_id=%s,api_type=%s,api_need_login=%s,api_is_remold=%s,api_status=%s where api_id =%s'
            params=(api_path,api_uri,method,case_desc,headers,data,str(group_id),str(api_type),str(api_need_login),str(api_is_remold),str(api_status),api_id)

        #db.exeUpdate(sql)
        db.exeUpdateByParamJson(sql, params)
        return_data = respdata().sucessMessage('', '更新成功')
        print(return_data)
        return json.dumps(return_data, ensure_ascii=False)


def loop_api(api_t):
    api_l = []
    for i in range(len(api_t)):
        api_l.append(api_t[i][0])
    return api_l
def nonexistent_api():
    return_data = respdata().failMessage('', '不存在的api')
    print(return_data)
    return json.dumps(return_data, ensure_ascii=False)