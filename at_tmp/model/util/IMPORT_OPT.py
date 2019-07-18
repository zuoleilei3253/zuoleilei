#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/7/18 14:15
# @Author  : bxf
# @File    : IMPORT_OPT.py
# @Software: PyCharm
from openpyxl import load_workbook
from model.util.PUB_LOG import *

class IMPORT_OPT:
    '''
    实现表格获取，行数获取
    '''
    def __init__(self,path):
        self.wb=load_workbook(path)
        exeLog("********加载Excel成功")

    def get_sheet(self,sheet_name):
        '''
        获取sheet_name页
        :param sheet_name: sheet名字
        :return:
        '''
        ws = self.wb.get_sheet_by_name(sheet_name)
        exeLog("********加载Sheet表成功")
        return ws
    def get_counts(self,sheet_name):
        '''
        获取行数
        :return:
        '''
        counts=self.get_sheet(sheet_name).max_row
        exeLog("********获取最大行数成功")
        return counts

